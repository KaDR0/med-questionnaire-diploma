import math
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from collections import defaultdict

from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from django.db import models, transaction
from openpyxl import load_workbook
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from .models import Answer, Assessment, DoctorProfile, LabIndicator, LabResult, LabValue, Patient, Question, Questionnaire
from .permissions import get_user_role


def _register_pdf_font():
    font_name = "Helvetica"
    project_root = Path(__file__).resolve().parents[1]
    candidate_paths = [
        project_root / "assets" / "fonts" / "DejaVuSans.ttf",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/local/share/fonts/DejaVuSans.ttf"),
        Path("/Library/Fonts/DejaVuSans.ttf"),
        Path("/System/Library/Fonts/Supplemental/Arial Unicode.ttf"),
    ]

    for font_path in candidate_paths:
        if font_path.exists():
            try:
                pdfmetrics.registerFont(TTFont("DejaVuSans", str(font_path)))
                font_name = "DejaVuSans"
                break
            except Exception:
                continue

    return font_name


def _calculate_answer_score(question, value):
    normalized_value = str(value).strip()

    if question.qtype == Question.YESNO:
        if normalized_value.lower() == "yes":
            return question.score_yes
        if normalized_value.lower() == "no":
            return question.score_no
    elif question.qtype == Question.SINGLE_CHOICE:
        for option in question.options or []:
            option_value = str(option.get("value", "")).strip()
            option_label = str(option.get("text", "")).strip()
            if normalized_value == option_value or normalized_value == option_label:
                try:
                    return int(option.get("score", 0))
                except (TypeError, ValueError):
                    return 0

    return 0


def _interpret_from_schema(total_score, max_score, completion_percent, questionnaire):
    schema = questionnaire.interpretation_schema or {}
    bands = schema.get("bands") or []
    default_schema = [
        {
            "min_ratio": 0.75,
            "label": "high_risk",
            "title": "High-risk screening result",
            "recommendation": "Clinical attention is recommended.",
        },
        {
            "min_ratio": 0.4,
            "label": "moderate_risk",
            "title": "Moderate-risk screening result",
            "recommendation": "Continue closer follow-up and repeat screening.",
        },
        {
            "min_ratio": 0.0,
            "label": "low_risk",
            "title": "Low-risk screening result",
            "recommendation": "Continue routine observation.",
        },
    ]

    active_bands = bands if bands else default_schema
    ratio = (total_score / max_score) if max_score > 0 else 0

    chosen_band = active_bands[-1]
    for band in active_bands:
        if ratio >= float(band.get("min_ratio", 0)):
            chosen_band = band
            break

    interpretation = {
        "label": chosen_band.get("label", "unclassified"),
        "title": chosen_band.get("title", "Screening result"),
        "recommendation": chosen_band.get("recommendation", "Review questionnaire answers."),
        "score_ratio": ratio,
        "target_condition_code": questionnaire.target_condition_code or questionnaire.disease.code,
        "questionnaire_kind": questionnaire.kind,
    }

    if completion_percent < questionnaire.min_completion_percent:
        interpretation["quality_note"] = "Low completion rate. Consider repeating this questionnaire."

    return interpretation


@transaction.atomic
def calculate_and_save_assessment(patient, questionnaire, answers, doctor=None):
    if isinstance(patient, int):
        patient = Patient.objects.get(id=patient)
    if isinstance(questionnaire, int):
        questionnaire = Questionnaire.objects.get(id=questionnaire)
    if isinstance(doctor, int):
        doctor = User.objects.filter(id=doctor).first()

    questions = list(Question.objects.filter(questionnaire=questionnaire).order_by("order"))
    assessment = Assessment.objects.create(
        patient=patient,
        questionnaire=questionnaire,
        doctor=doctor,
        total_score=0,
        completion_percent=0,
        quality_flag=Assessment.QUALITY_VALID,
        interpretation={},
        conclusion="",
    )

    total_score = 0
    answered_count = 0
    max_score = 0
    feature_values = {}

    for question in questions:
        raw_value = answers.get(str(question.id), answers.get(question.id, ""))
        value = str(raw_value).strip()

        if value:
            answered_count += 1

        score = _calculate_answer_score(question, value)

        if question.qtype == Question.YESNO:
            max_score += max(question.score_yes, question.score_no)
        elif question.qtype == Question.SINGLE_CHOICE:
            options = question.options or []
            max_option_score = 0
            for option in options:
                try:
                    max_option_score = max(max_option_score, int(option.get("score", 0)))
                except (TypeError, ValueError):
                    continue
            max_score += max_option_score

        Answer.objects.create(
            assessment=assessment,
            question=question,
            value=value,
            score=score,
        )
        total_score += score
        if question.feature_key:
            normalized = value.lower()
            if normalized in {"yes", "true", "1"}:
                feature_values[question.feature_key] = 1
            elif normalized in {"no", "false", "0"}:
                feature_values[question.feature_key] = 0
            else:
                feature_values[question.feature_key] = value

    total_questions = len(questions)
    completion_percent = (answered_count / total_questions * 100) if total_questions else 0
    quality_flag = Assessment.QUALITY_VALID
    if completion_percent < questionnaire.min_completion_percent:
        quality_flag = Assessment.QUALITY_PARTIAL if answered_count > 0 else Assessment.QUALITY_INVALID

    interpretation = _interpret_from_schema(total_score, max_score, completion_percent, questionnaire)
    interpretation["visit_summary"] = {
        "questionnaire_title": questionnaire.title_en or questionnaire.title_ru or questionnaire.title_kk or questionnaire.title,
        "score": total_score,
        "risk_level": interpretation.get("label", "unclassified"),
        "quality_flag": quality_flag,
        "completion_percent": round(completion_percent, 2),
        "key_features": feature_values,
    }
    conclusion = f"{interpretation['title']}. {interpretation['recommendation']}"

    assessment.total_score = total_score
    assessment.completion_percent = round(completion_percent, 2)
    assessment.quality_flag = quality_flag
    assessment.interpretation = interpretation
    assessment.conclusion = conclusion
    assessment.save(
        update_fields=["total_score", "completion_percent", "quality_flag", "interpretation", "conclusion"]
    )

    return assessment


def _normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def _to_int(value):
    value = _normalize_cell(value)
    if value is None:
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_float(value):
    value = _normalize_cell(value)
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_indicator_name(raw_name):
    if raw_name is None:
        return ""
    return str(raw_name).strip().lower()


def _resolve_indicator_by_name(raw_name, cache):
    key = _normalize_indicator_name(raw_name)
    if not key:
        return None
    return cache.get(key)


def _build_indicator_lookup():
    lookup = {}
    for indicator in LabIndicator.objects.all():
        names = {indicator.name, indicator.standard_name, indicator.code}
        for item in (indicator.synonyms or []):
            names.add(item)
        for name in names:
            normalized = _normalize_indicator_name(name)
            if normalized:
                lookup[normalized] = indicator
    return lookup


def _build_patient_name(row_data, created_index):
    for field in ["full_name", "Full Name", "name", "Name", "ФИО", "fio"]:
        value = _normalize_cell(row_data.get(field))
        if value:
            return str(value)
    patient_code = _normalize_cell(row_data.get("Patient_ID"))
    if patient_code:
        return f"Patient {patient_code}"
    return f"Imported Patient {created_index}"


@transaction.atomic
def import_patients_from_excel(file_obj, importing_user=None):
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {
            "created": 0,
            "skipped": 0,
            "errors": ["The Excel file is empty."],
            "imported_patients": [],
        }

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if "Patient_ID" not in headers:
        raise ValueError("Patient_ID column is required in the Excel file.")

    created = 0
    skipped = 0
    errors = []
    imported_patients = []

    for row_number, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_data = {
            headers[index]: _normalize_cell(row[index]) if index < len(row) else None
            for index in range(len(headers))
            if headers[index]
        }

        patient_code = _normalize_cell(row_data.get("Patient_ID"))
        if not patient_code:
            errors.append(f"Row {row_number}: Patient_ID is empty.")
            continue

        patient_code = str(patient_code)
        if Patient.objects.filter(patient_code=patient_code).exists():
            skipped += 1
            continue

        age = _to_int(row_data.get("age"))
        sex = _to_int(row_data.get("sex 1-men.2-women") or row_data.get("sex"))
        height_cm = _to_float(row_data.get("h(sm)") or row_data.get("height_cm"))
        weight_kg = _to_float(row_data.get("m") or row_data.get("weight_kg"))

        patient = Patient.objects.create(
            patient_code=patient_code,
            full_name=_build_patient_name(row_data, created + 1),
            age=age,
            sex=sex,
            height_cm=height_cm,
            weight_kg=weight_kg,
            data=row_data,
            created_by=importing_user if getattr(importing_user, "is_authenticated", False) else None,
            updated_by=importing_user if getattr(importing_user, "is_authenticated", False) else None,
            assigned_doctor=(
                importing_user
                if getattr(importing_user, "is_authenticated", False)
                and get_user_role(importing_user) == DoctorProfile.ROLE_DOCTOR
                else None
            ),
        )
        created += 1
        imported_patients.append(
            {
                "id": patient.id,
                "patient_code": patient.patient_code,
                "full_name": patient.full_name,
            }
        )

    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "imported_patients": imported_patients,
    }


def _parse_excel_date(value):
    value = _normalize_cell(value)
    if value is None:
        return date.today()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return date.today()


def create_lab_template_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Lab Import"

    headers = ["Patient_ID", "Date"] + list(LabIndicator.objects.order_by("name").values_list("name", flat=True))
    worksheet.append(headers)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def create_patient_template_workbook():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Patient Import"

    headers = ["Patient_ID", "full_name", "age", "sex", "height_cm", "weight_kg"]
    worksheet.append(headers)
    worksheet.append(["000000000001", "John Doe", 45, 1, 178, 82])
    worksheet.append(["000000000002", "Aigerim N.", 34, 2, 165, 61])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@transaction.atomic
def import_labs_from_excel(file_obj, importing_user=None):
    workbook = load_workbook(file_obj, data_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return {
            "created_results": 0,
            "created_values": 0,
            "skipped": 0,
            "errors": ["The Excel file is empty."],
            "affected_patient_ids": [],
        }

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    if "Patient_ID" not in headers:
        raise ValueError("Patient_ID column is required in the Excel file.")

    indicators = _build_indicator_lookup()
    if not indicators:
        raise ValueError("No lab indicators found. Run seed_lab_indicators first.")

    created_results = 0
    created_values = 0
    skipped = 0
    errors = []
    touched_patient_ids = set()

    long_format_keys = {
        "analysis_name",
        "analysis",
        "test_name",
        "indicator",
        "value",
        "unit",
        "ref_min",
        "ref_max",
        "date",
        "patient_id",
    }
    headers_normalized = {_normalize_indicator_name(h) for h in headers}
    is_long_format = len(headers_normalized.intersection(long_format_keys)) >= 4

    role = get_user_role(importing_user) if getattr(importing_user, "is_authenticated", False) else None

    if is_long_format:
        grouped_rows = defaultdict(list)
        all_rows = list(rows)
        for row_number, row in enumerate(all_rows, start=2):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_data = {
                headers[index]: _normalize_cell(row[index]) if index < len(row) else None
                for index in range(len(headers))
                if headers[index]
            }
            patient_code = _normalize_cell(
                row_data.get("Patient_ID")
                or row_data.get("patient_id")
                or row_data.get("patient")
            )
            if not patient_code:
                errors.append(f"Row {row_number}: Patient_ID is empty.")
                continue
            grouped_rows[str(patient_code)].append((row_number, row_data))

        for patient_code, patient_rows in grouped_rows.items():
            patient_queryset = Patient.objects.filter(patient_code=str(patient_code))
            if role == DoctorProfile.ROLE_DOCTOR:
                patient_queryset = patient_queryset.filter(
                    models.Q(assigned_doctor=importing_user) | models.Q(created_by=importing_user)
                )
            patient = patient_queryset.first()
            if not patient:
                skipped += len(patient_rows)
                if role == DoctorProfile.ROLE_DOCTOR:
                    errors.append(
                        f"Patient with Patient_ID {patient_code} was not found or is not assigned to current doctor."
                    )
                else:
                    errors.append(f"Patient with Patient_ID {patient_code} was not found.")
                continue

            by_date = defaultdict(list)
            for row_number, row_data in patient_rows:
                row_date = _parse_excel_date(row_data.get("Date") or row_data.get("date"))
                by_date[row_date].append((row_number, row_data))

            for row_date, rows_for_date in by_date.items():
                lab_result = LabResult.objects.create(patient=patient, date=row_date)
                row_created_values = 0
                for row_number, row_data in rows_for_date:
                    indicator_name = (
                        row_data.get("analysis_name")
                        or row_data.get("analysis")
                        or row_data.get("test_name")
                        or row_data.get("indicator")
                        or row_data.get("name")
                    )
                    indicator = _resolve_indicator_by_name(indicator_name, indicators)
                    numeric_value = _to_float(row_data.get("value") or row_data.get("result"))

                    if not indicator or numeric_value is None:
                        continue

                    if _normalize_cell(row_data.get("unit")) and not indicator.unit:
                        indicator.unit = str(_normalize_cell(row_data.get("unit")))
                    if _to_float(row_data.get("ref_min")) is not None:
                        indicator.min_norm = _to_float(row_data.get("ref_min"))
                    if _to_float(row_data.get("ref_max")) is not None:
                        indicator.max_norm = _to_float(row_data.get("ref_max"))
                    indicator.save(update_fields=["unit", "min_norm", "max_norm"])

                    LabValue.objects.create(
                        result=lab_result,
                        indicator=indicator,
                        value=numeric_value,
                    )
                    row_created_values += 1
                    created_values += 1

                if row_created_values == 0:
                    lab_result.delete()
                    skipped += 1
                    errors.append(f"Date {row_date}: no matching lab values were found.")
                    continue
                created_results += 1
                touched_patient_ids.add(patient.id)

        return {
            "created_results": created_results,
            "created_values": created_values,
            "skipped": skipped,
            "errors": errors,
            "affected_patient_ids": list(touched_patient_ids),
        }

    for row_number, row in enumerate(rows, start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        row_data = {
            headers[index]: _normalize_cell(row[index]) if index < len(row) else None
            for index in range(len(headers))
            if headers[index]
        }

        patient_code = _normalize_cell(row_data.get("Patient_ID"))
        if not patient_code:
            errors.append(f"Row {row_number}: Patient_ID is empty.")
            continue

        patient_queryset = Patient.objects.filter(patient_code=str(patient_code))
        if role == DoctorProfile.ROLE_DOCTOR:
            patient_queryset = patient_queryset.filter(
                models.Q(assigned_doctor=importing_user) | models.Q(created_by=importing_user)
            )
        patient = patient_queryset.first()
        if not patient:
            skipped += 1
            if role == DoctorProfile.ROLE_DOCTOR:
                errors.append(
                    f"Row {row_number}: patient with Patient_ID {patient_code} was not found or is not assigned to current doctor."
                )
            else:
                errors.append(f"Row {row_number}: patient with Patient_ID {patient_code} was not found.")
            continue

        result_date = _parse_excel_date(row_data.get("Date"))
        lab_result = LabResult.objects.create(patient=patient, date=result_date)
        row_created_values = 0

        for header, raw_value in row_data.items():
            if header in {"Patient_ID", "Date"}:
                continue
            indicator = _resolve_indicator_by_name(header, indicators)
            numeric_value = _to_float(raw_value)
            if not indicator or numeric_value is None:
                continue

            LabValue.objects.create(
                result=lab_result,
                indicator=indicator,
                value=numeric_value,
            )
            row_created_values += 1
            created_values += 1

        if row_created_values == 0:
            lab_result.delete()
            skipped += 1
            errors.append(f"Row {row_number}: no matching lab values were found.")
            continue

        created_results += 1
        touched_patient_ids.add(patient.id)

    return {
        "created_results": created_results,
        "created_values": created_values,
        "skipped": skipped,
        "errors": errors,
        "affected_patient_ids": list(touched_patient_ids),
    }


def build_assessment_pdf(assessment):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    font_name = _register_pdf_font()
    for style_name in ["Title", "Heading2", "Heading3", "Normal", "BodyText"]:
        styles[style_name].fontName = font_name
    story = []

    patient = assessment.patient
    questionnaire = assessment.questionnaire
    doctor_name = ""
    if assessment.doctor:
        doctor_name = f"{assessment.doctor.first_name} {assessment.doctor.last_name}".strip() or assessment.doctor.username

    title = questionnaire.title_en or questionnaire.title_ru or questionnaire.title_kk or questionnaire.title
    story.append(Paragraph("Assessment Report", styles["Title"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Patient: {patient.full_name}", styles["Normal"]))
    story.append(Paragraph(f"Patient ID: {patient.patient_code or '-'}", styles["Normal"]))
    story.append(Paragraph(f"Questionnaire: {title or '-'}", styles["Normal"]))
    story.append(Paragraph(f"Doctor: {doctor_name or '-'}", styles["Normal"]))
    story.append(Paragraph(f"Created at: {assessment.created_at.strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 10))

    summary_data = [
        ["Total score", str(assessment.total_score)],
        ["Conclusion", assessment.conclusion or "-"],
    ]
    summary_table = Table(summary_data, colWidths=[45 * mm, 120 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(summary_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Submitted Answers", styles["Heading2"]))
    answer_rows = [["#", "Question", "Answer", "Score"]]
    for index, answer in enumerate(assessment.answers.select_related("question").all(), start=1):
        question_text = (
            answer.question.text_en
            or answer.question.text_ru
            or answer.question.text_kk
            or answer.question.text
        )
        answer_rows.append([str(index), question_text or "-", answer.value or "-", str(answer.score)])

    answer_table = Table(answer_rows, colWidths=[10 * mm, 90 * mm, 55 * mm, 20 * mm], repeatRows=1)
    answer_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("PADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(answer_table)

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_patient_report_pdf(patient):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    font_name = _register_pdf_font()
    for style_name in ["Title", "Heading2", "Heading3", "Normal", "BodyText"]:
        styles[style_name].fontName = font_name
    story = []

    sex_label = {1: "Male", 2: "Female"}.get(patient.sex, "-")
    status_label = patient.get_status_display() if patient.status else "-"
    next_visit = patient.next_visit_date.isoformat() if patient.next_visit_date else "-"

    story.append(Paragraph("Patient Report", styles["Title"]))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Patient: {patient.full_name}", styles["Normal"]))
    story.append(Paragraph(f"Patient ID: {patient.patient_code or '-'}", styles["Normal"]))
    story.append(Paragraph(f"Status: {status_label}", styles["Normal"]))
    story.append(Paragraph(f"Next visit: {next_visit}", styles["Normal"]))
    story.append(Spacer(1, 10))

    overview_rows = [
        ["Age", str(patient.age) if patient.age is not None else "-"],
        ["Sex", sex_label],
        ["Height", f"{patient.height_cm} cm" if patient.height_cm is not None else "-"],
        ["Weight", f"{patient.weight_kg} kg" if patient.weight_kg is not None else "-"],
    ]
    overview_table = Table(overview_rows, colWidths=[40 * mm, 120 * mm])
    overview_table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("PADDING", (0, 0), (-1, -1), 6),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    story.append(overview_table)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Assessment History", styles["Heading2"]))
    assessments = list(patient.assessments.select_related("questionnaire").order_by("-created_at")[:10])
    if assessments:
        assessment_rows = [["Date", "Questionnaire", "Score", "Conclusion"]]
        for assessment in assessments:
            questionnaire_title = (
                assessment.questionnaire.title_en
                or assessment.questionnaire.title_ru
                or assessment.questionnaire.title_kk
                or assessment.questionnaire.title
            )
            assessment_rows.append(
                [
                    assessment.created_at.strftime("%Y-%m-%d %H:%M"),
                    questionnaire_title or "-",
                    str(assessment.total_score),
                    assessment.conclusion or "-",
                ]
            )
        assessment_table = Table(
            assessment_rows,
            colWidths=[30 * mm, 65 * mm, 18 * mm, 55 * mm],
            repeatRows=1,
        )
        assessment_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("PADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(assessment_table)
    else:
        story.append(Paragraph("No assessments available.", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Questionnaire Answers (latest assessments)", styles["Heading2"]))
    detailed_assessments = (
        patient.assessments.select_related("questionnaire")
        .prefetch_related("answers__question")
        .order_by("-created_at")[:3]
    )
    if detailed_assessments:
        for assessment in detailed_assessments:
            questionnaire_title = (
                assessment.questionnaire.title_en
                or assessment.questionnaire.title_ru
                or assessment.questionnaire.title_kk
                or assessment.questionnaire.title
            )
            story.append(
                Paragraph(
                    f"{assessment.created_at.strftime('%Y-%m-%d %H:%M')} - {questionnaire_title or '-'}",
                    styles["Heading3"],
                )
            )
            answer_rows = [["Question", "Answer", "Score"]]
            for answer in assessment.answers.all():
                question_text = (
                    answer.question.text_en
                    or answer.question.text_ru
                    or answer.question.text_kk
                    or answer.question.text
                )
                answer_rows.append([question_text or "-", answer.value or "-", str(answer.score)])
            answer_table = Table(answer_rows, colWidths=[95 * mm, 55 * mm, 20 * mm], repeatRows=1)
            answer_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                        ("PADDING", (0, 0), (-1, -1), 5),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ]
                )
            )
            story.append(answer_table)
            story.append(Spacer(1, 6))
    else:
        story.append(Paragraph("No questionnaire answers available.", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Laboratory Results", styles["Heading2"]))
    lab_results = list(patient.labresult_set.prefetch_related("values__indicator").order_by("-date", "-id")[:5])
    if lab_results:
        for lab_result in lab_results:
            story.append(Paragraph(f"Date: {lab_result.date.isoformat()}", styles["Normal"]))
            lab_rows = [["Indicator", "Value", "Norm"]]
            for value in lab_result.values.all():
                norm = f"{value.indicator.min_norm if value.indicator.min_norm is not None else '-'} - {value.indicator.max_norm if value.indicator.max_norm is not None else '-'} {value.indicator.unit or ''}".strip()
                lab_rows.append(
                    [
                        value.indicator.name,
                        f"{value.value} {value.indicator.unit or ''}".strip(),
                        norm,
                    ]
                )
            lab_table = Table(lab_rows, colWidths=[60 * mm, 45 * mm, 55 * mm], repeatRows=1)
            lab_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ecfeff")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                        ("PADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            story.append(lab_table)
            story.append(Spacer(1, 8))
    else:
        story.append(Paragraph("No laboratory results available.", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Clinical Risk Profile", styles["Heading2"]))
    latest_risk_profile = (
        patient.risk_profiles.prefetch_related("findings__recommendation_template", "red_flags")
        .order_by("-created_at")
        .first()
    )
    if latest_risk_profile:
        story.append(
            Paragraph(
                f"Generated at: {latest_risk_profile.created_at.strftime('%Y-%m-%d %H:%M')} | "
                f"Overall risk level: {latest_risk_profile.overall_risk_level}",
                styles["Normal"],
            )
        )
        story.append(Paragraph(latest_risk_profile.summary or "-", styles["BodyText"]))
        story.append(Spacer(1, 6))

        findings = list(latest_risk_profile.findings.all())
        if findings:
            story.append(Paragraph("Detected risk findings", styles["Heading3"]))
            finding_rows = [["Problem", "Risk", "Evidence", "ML"]]
            for finding in findings:
                evidence_text = "; ".join(finding.evidence or []) or "-"
                ml_text = "-"
                if finding.ml_probability is not None:
                    ml_text = f"{round(finding.ml_probability * 100)}%"
                    if finding.confidence_score is not None:
                        ml_text += f" (conf {round(finding.confidence_score * 100)}%)"
                finding_rows.append(
                    [
                        finding.problem_code,
                        finding.risk_level,
                        evidence_text,
                        ml_text,
                    ]
                )
            findings_table = Table(
                finding_rows,
                colWidths=[35 * mm, 20 * mm, 90 * mm, 30 * mm],
                repeatRows=1,
            )
            findings_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                        ("PADDING", (0, 0), (-1, -1), 5),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(findings_table)
            story.append(Spacer(1, 6))

            story.append(Paragraph("Recommendations", styles["Heading3"]))
            for idx, finding in enumerate(findings, start=1):
                tpl = finding.recommendation_template
                if not tpl:
                    continue
                story.append(
                    Paragraph(
                        f"{idx}. {finding.problem_code}: {tpl.preliminary_conclusion} {tpl.next_steps}",
                        styles["BodyText"],
                    )
                )
            story.append(Spacer(1, 6))

        red_flags = list(latest_risk_profile.red_flags.all())
        if red_flags:
            story.append(Paragraph("Red flags", styles["Heading3"]))
            red_flag_rows = [["Urgency", "Trigger signs", "Recommended action"]]
            for flag in red_flags:
                red_flag_rows.append(
                    [
                        flag.urgency_level,
                        ", ".join(flag.trigger_signs or []) or "-",
                        flag.recommended_action or "-",
                    ]
                )
            red_flags_table = Table(red_flag_rows, colWidths=[25 * mm, 70 * mm, 80 * mm], repeatRows=1)
            red_flags_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fee2e2")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                        ("PADDING", (0, 0), (-1, -1), 5),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(red_flags_table)
        else:
            story.append(Paragraph("No red flags detected in the latest profile.", styles["BodyText"]))

        history_profiles = list(
            patient.risk_profiles.prefetch_related("findings", "red_flags")
            .order_by("-created_at")[:6]
        )
        if history_profiles:
            story.append(Spacer(1, 6))
            story.append(Paragraph("Risk dynamics (recent history)", styles["Heading3"]))
            history_rows = [["Date", "Overall risk", "Findings", "Red flags"]]
            for profile in history_profiles:
                history_rows.append(
                    [
                        profile.created_at.strftime("%Y-%m-%d %H:%M"),
                        profile.overall_risk_level,
                        str(profile.findings.count()),
                        str(profile.red_flags.count()),
                    ]
                )
            history_table = Table(history_rows, colWidths=[40 * mm, 35 * mm, 30 * mm, 30 * mm], repeatRows=1)
            history_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef2ff")),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                        ("PADDING", (0, 0), (-1, -1), 5),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]
                )
            )
            story.append(history_table)
    else:
        story.append(Paragraph("No risk profile available.", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Doctor Notes", styles["Heading2"]))
    notes = list(patient.notes.select_related("doctor").order_by("-pinned", "-created_at")[:8])
    if notes:
        for note in notes:
            doctor_name = "-"
            if note.doctor:
                doctor_name = f"{note.doctor.first_name} {note.doctor.last_name}".strip() or note.doctor.username
            note_title = f"{'[Pinned] ' if note.pinned else ''}{note.get_category_display()} - {doctor_name}"
            story.append(Paragraph(note_title, styles["Heading3"]))
            story.append(Paragraph(note.text or "-", styles["Normal"]))
            story.append(Paragraph(note.created_at.strftime("%Y-%m-%d %H:%M"), styles["BodyText"]))
            story.append(Spacer(1, 6))
    else:
        story.append(Paragraph("No doctor notes available.", styles["Normal"]))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Imported Excel Data", styles["Heading2"]))
    excel_rows = [["Field", "Value"]]
    for key, value in list((patient.data or {}).items())[:30]:
        excel_rows.append([str(key), "-" if value in (None, "") else str(value)])
    if len(excel_rows) > 1:
        excel_table = Table(excel_rows, colWidths=[65 * mm, 95 * mm], repeatRows=1)
        excel_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f8fafc")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("PADDING", (0, 0), (-1, -1), 5),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]
            )
        )
        story.append(excel_table)
    else:
        story.append(Paragraph("No imported Excel data available.", styles["Normal"]))

    from xml.sax.saxutils import escape

    story.append(Spacer(1, 16))
    story.append(Paragraph("Doctor's orders / Назначение врача", styles["Heading2"]))
    story.append(Spacer(1, 6))
    try:
        raw_order = (patient.doctor_order.order_text or "").strip()
    except ObjectDoesNotExist:
        raw_order = ""

    if not raw_order:
        story.append(Paragraph("No doctor's orders recorded.", styles["Normal"]))
    else:
        for line in raw_order.splitlines():
            story.append(Paragraph(escape(line) if line.strip() else " ", styles["BodyText"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def build_doctor_order_pdf(patient, order_text: str):
    from xml.sax.saxutils import escape

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    font_name = _register_pdf_font()
    for style_name in ["Title", "Heading2", "Normal", "BodyText"]:
        styles[style_name].fontName = font_name

    story = []
    sex_label = {1: "Male", 2: "Female"}.get(patient.sex, "—")
    story.append(Paragraph("Doctor's orders / Назначение врача", styles["Title"]))
    story.append(Spacer(1, 10))
    story.append(Paragraph("<b>Patient</b>", styles["Heading2"]))
    story.append(Paragraph(escape(f"Name: {patient.full_name}"), styles["Normal"]))
    story.append(Paragraph(escape(f"Patient ID: {patient.patient_code or '—'}"), styles["Normal"]))
    story.append(Paragraph(escape(f"Age: {patient.age if patient.age is not None else '—'}"), styles["Normal"]))
    story.append(Paragraph(escape(f"Sex: {sex_label}"), styles["Normal"]))
    nv = patient.next_visit_date.isoformat() if patient.next_visit_date else "—"
    story.append(Paragraph(escape(f"Next visit: {nv}"), styles["Normal"]))
    story.append(Spacer(1, 14))
    story.append(Paragraph("<b>Individual prescriptions (free text)</b>", styles["Heading2"]))
    story.append(Spacer(1, 6))

    body = (order_text or "").strip() or "—"
    for line in body.splitlines():
        story.append(Paragraph(escape(line) if line.strip() else " ", styles["BodyText"]))
    story.append(Spacer(1, 24))
    story.append(
        Paragraph(
            "<i>This document is a template for recording the physician's orders. "
            "It supplements, but does not replace, the medical records kept at the facility.</i>",
            styles["Normal"],
        )
    )

    doc.build(story)
    buffer.seek(0)
    return buffer
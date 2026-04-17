from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from django.contrib import messages
from django.views.decorators.http import require_POST
import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import json
from django.utils.safestring import mark_safe
from .models import PatientNote
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors

from django.http import HttpResponse
from .models import LabIndicator
import openpyxl
from .models import LabIndicator, LabResult, LabValue
from django.views.decorators.http import require_POST
from .models import Assessment
from .models import Patient, Questionnaire, Question
from .services import calculate_and_save_assessment

from django.http import HttpResponse
from .models import LabIndicator
import openpyxl
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics import renderPDF
from reportlab.lib import colors
from .models import LabResult, LabValue

from django.http import HttpResponse
from django.utils import timezone

from django.conf import settings
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from .models import Answer
from datetime import datetime, date
from django.contrib import messages
from django.views.decorators.http import require_POST
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics import renderPDF

def draw_score_chart_pdf(c, labels, scores, x=40, y=650, w=515, h=140):
    """
    c = reportlab canvas
    labels = ["14.02.2026 01:11", ...]
    scores = [1, 3, 2, ...]
    x,y,w,h - позиция и размер на странице
    """
    if not scores:
        return

    d = Drawing(w, h)
    chart = HorizontalLineChart()
    chart.x = 40
    chart.y = 20
    chart.width = w - 60
    chart.height = h - 40

    # данные
    chart.data = [scores]

    # оси
    chart.valueAxis.valueMin = 0
    chart.valueAxis.valueMax = max(scores) + 1

    # X-ось как категории (даты)
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.boxAnchor = "ne"
    chart.categoryAxis.labels.angle = 45
    chart.categoryAxis.labels.dx = -10
    chart.categoryAxis.labels.dy = -5

    d.add(chart)
    renderPDF.draw(d, c, x, y)

@login_required
def patient_list(request):
    query = request.GET.get("q", "").strip()

    patients = Patient.objects.all()

    if query:
        patients = patients.filter(patient_code__icontains=query)

    return render(request, "core/patient_list.html", {
        "patients": patients,
        "query": query,
    })

@login_required
def start_assessment(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    questionnaires = Questionnaire.objects.filter(is_active=True)
    return render(
        request,
        "core/start_assessment.html",
        {"patient": patient, "questionnaires": questionnaires},
    )


@login_required
def questionnaire_view(request, patient_id, questionnaire_id):
    patient = get_object_or_404(Patient, id=patient_id)
    questionnaire = get_object_or_404(Questionnaire, id=questionnaire_id)
    questions = Question.objects.filter(questionnaire=questionnaire)

    if request.method == "POST":
        answers = dict(request.POST)  # ключи = id вопросов
        assessment = calculate_and_save_assessment(
            patient=patient,
            questionnaire=questionnaire,
            doctor=request.user,
            answers_dict=answers,
        )
        # пока редирект можно временно убрать, если результата еще нет
        return redirect("assessment_result", assessment.id)

    return render(
        request,
        "core/questionnaire.html",
        {
            "patient": patient,
            "questionnaire": questionnaire,
            "questions": questions,
        },
    )
@login_required
def patient_detail(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)

    selected_analysis = request.GET.get("analysis", "").strip()

    history = Assessment.objects.filter(patient=patient).order_by("created_at")
    labels = [h.created_at.strftime("%d.%m.%Y %H:%M") for h in history]
    scores = [h.total_score for h in history]
    notes = patient.notes.all().order_by("-created_at")
    
    lab_results = (
        LabResult.objects
        .filter(patient=patient)
        .prefetch_related("values__indicator")
        .order_by("date")
    )

    # 📋 список доступных анализов
    available_analyses = set()
    for result in lab_results:
        for val in result.values.all():
            available_analyses.add(val.indicator.name)

    available_analyses = sorted(available_analyses)

    # 📊 графики
    lab_charts = {}

    for result in lab_results:
        for val in result.values.all():
            name = val.indicator.name

            # 🔥 фильтр
            if selected_analysis and name != selected_analysis:
                continue

            if name not in lab_charts:
                lab_charts[name] = {
                    "labels": [],
                    "values": []
                }

            lab_charts[name]["labels"].append(result.date.strftime("%d.%m.%Y"))
            lab_charts[name]["values"].append(val.value)

    return render(request, "core/patient_detail.html", {
        "patient": patient,
        "history": history,
        "labels_json": mark_safe(json.dumps(labels)),
        "scores_json": mark_safe(json.dumps(scores)),
        "lab_results": lab_results,
        "lab_charts_json": mark_safe(json.dumps(lab_charts)),
        "available_analyses": available_analyses,
        "selected_analysis": selected_analysis,
        "notes": notes,
    })

    # Данные для графиков анализов
    lab_charts = {}

    for result in lab_results:
        for val in result.values.all():
            name = val.indicator.name

            if name not in lab_charts:
                lab_charts[name] = {
                    "labels": [],
                    "values": []
                }

            lab_charts[name]["labels"].append(result.date.strftime("%d.%m.%Y"))
            lab_charts[name]["values"].append(val.value)

    return render(request, "core/patient_detail.html", {
        "patient": patient,
        "history": history,
        "labels_json": mark_safe(json.dumps(labels)),
        "scores_json": mark_safe(json.dumps(scores)),
        "lab_results": lab_results,
        "lab_charts_json": mark_safe(json.dumps(lab_charts)),
    })

@login_required
def assessment_result(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    answers = assessment.answers.select_related("question").all()
    return render(request, "core/assessment_result.html", {
        "assessment": assessment,
        "answers": answers,
    })

@login_required
def assessment_pdf(request, assessment_id):
    assessment = get_object_or_404(Assessment, id=assessment_id)
    patient = assessment.patient
    answers = (
        Answer.objects
        .filter(assessment=assessment)
        .select_related("question")
        .order_by("question__order")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="assessment_{assessment.id}.pdf"'

    c = canvas.Canvas(response, pagesize=A4)

    font_path = os.path.join("fonts", "DejaVuSans.ttf")
    pdfmetrics.registerFont(TTFont("DejaVu", font_path))
    c.setFont("DejaVu", 11)

    width, height = A4
    y = height - 50

    def line(text="", step=16):
        nonlocal y
        c.setFont("DejaVu", 11)  # <-- ставим шрифт каждый раз
        c.drawString(40, y, str(text)[:110])
        y -= step
        if y < 60:
            c.showPage()
            c.setFont("DejaVu", 11)  # <-- ВАЖНО после новой страницы
            y = height - 50

    # Заголовок
    line("Medical Questionnaire Report", step=22)
    line(f"Assessment ID: {assessment.id}")
    line(f"Date/Time (Almaty): {timezone.localtime(assessment.created_at).strftime('%d.%m.%Y %H:%M')}")
    line("")

    # Пациент
    line("Patient:", step=18)
    line(f"Full name: {patient.full_name}")
    if hasattr(patient, "age"):
        line(f"Age: {patient.age}")
    line(f"Sex (1-men,2-women): {patient.sex}")
    line(f"Height (cm): {patient.height_cm}")
    line(f"Weight (kg): {patient.weight_kg}")
    line("")

    # Все данные из Excel (JSON)
    if hasattr(patient, "data") and isinstance(patient.data, dict):
        line("All data from Excel:", step=18)
        for k, v in patient.data.items():
            line(f"{k}: {v}")
        line("")

    # Опросник + итог
    line("Questionnaire:", step=18)
    line(f"Title: {assessment.questionnaire.title}")
    line(f"Total score: {assessment.total_score}")
    line(f"Conclusion: {assessment.conclusion}")
    line("")

    # Ответы
    line("Answers:", step=18)
    for a in answers:
        q = a.question
        line(f"{q.order}. {q.text}")
        line(f"   Answer: {a.value}   Score: {a.score}")

    c.showPage()
    c.save()
    return response
from .models import Answer  # если нет в импортах

@login_required
def patient_pdf_all(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)

    assessments = (
        Assessment.objects
        .filter(patient=patient)
        .select_related("questionnaire")
        .order_by("created_at")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="patient_{patient.id}_all_assessments.pdf"'

    c = canvas.Canvas(response, pagesize=A4)

    font_path = os.path.join("fonts", "DejaVuSans.ttf")
    pdfmetrics.registerFont(TTFont("DejaVu", font_path))
    c.setFont("DejaVu", 11)

    width, height = A4
    y = height - 50

    def line(text="", step=16):
        nonlocal y
        c.setFont("DejaVu", 11)
        c.drawString(40, y, str(text)[:110])
        y -= step
        if y < 60:
            c.showPage()
            c.setFont("DejaVu", 11)
            y = height - 50

    # Заголовок
    line("Patient Full Report (All Questionnaires)", step=22)
    line(f"Generated (Almaty): {timezone.localtime(timezone.now()).strftime('%d.%m.%Y %H:%M')}")
    line("")

    # Пациент
    line("Patient:", step=18)
    line(f"Full name: {patient.full_name}")
    if hasattr(patient, "age"):
        line(f"Age: {patient.age}")
    line(f"Sex: {patient.sex}")
    line(f"Height (cm): {patient.height_cm}")
    line(f"Weight (kg): {patient.weight_kg}")
    line("")
# ===== АНАЛИЗЫ ПАЦИЕНТА =====
    lab_results = (
        LabResult.objects
        .filter(patient=patient)
        .prefetch_related("values__indicator")
        .order_by("-date")
    )

    if lab_results.exists():
        line("Lab Results:", step=18)

        for lab_result in lab_results:
            line(f"Date: {lab_result.date.strftime('%d.%m.%Y')}", step=18)

            table_data = [["Indicator", "Value", "Norm", "Status"]]

            row_styles = []

            for val in lab_result.values.all():
                min_norm = val.indicator.min_norm
                max_norm = val.indicator.max_norm

                if min_norm is not None and val.value < min_norm:
                    status = "Low"
                    status_color = colors.orange
                elif max_norm is not None and val.value > max_norm:
                    status = "High"
                    status_color = colors.red
                else:
                    status = "Normal"
                    status_color = colors.green

                norm_text = f"{min_norm if min_norm is not None else '?'} - {max_norm if max_norm is not None else '?'}"
                value_text = f"{val.value} {val.indicator.unit or ''}".strip()

                table_data.append([
                    val.indicator.name,
                    value_text,
                    norm_text,
                    status
                ])

                row_index = len(table_data) - 1
                row_styles.append(("TEXTCOLOR", (3, row_index), (3, row_index), status_color))

            table = Table(table_data, colWidths=[160, 100, 100, 100])


            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (-1, -1), "DejaVu"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ] + row_styles))

            table_width, table_height = table.wrap(500, 700)

            if y - table_height < 60:
                c.showPage()
                c.setFont("DejaVu", 11)
                y = height - 50

            table.drawOn(c, 40, y - table_height)
            y -= table_height + 20

    # ===== ГРАФИК ОДНОГО АНАЛИЗА (например Iron) =====
    iron_data = []

    for a in lab_results:
        for v in a.values.all():
            if v.indicator.name.lower() == "iron":
                dt = timezone.localtime(a.created_at)
                iron_data.append((dt, v.value))

    if iron_data:
        line("Динамика Iron:", step=18)

        drawing = Drawing(400, 200)
        chart = LinePlot()

        chart.x = 50
        chart.y = 40
        chart.height = 120
        chart.width = 300

        # преобразуем даты в числа (индексы)
        data = [(i, val[1]) for i, val in enumerate(iron_data)]
        chart.data = [data]

        chart.lines[0].strokeColor = colors.blue
        chart.lines[0].strokeWidth = 2

        chart.xValueAxis.valueMin = 0
        chart.xValueAxis.valueMax = len(data)

        chart.yValueAxis.valueMin = 0
        chart.yValueAxis.valueMax = max([v[1] for v in iron_data]) + 5

        drawing.add(chart)

        renderPDF.draw(drawing, c, 40, y - 180)
        y -= 200

        line("")
# ===== ГРАФИК ДИНАМИКИ БАЛЛОВ =====
    history = (
        Assessment.objects
        .filter(patient=patient)
        .order_by("created_at")
    )

    if history.exists():
        line("Score dynamics:", step=18)

        labels = []
        scores = []

        for a in history:
            labels.append(timezone.localtime(a.created_at).strftime("%d.%m"))
            scores.append(a.total_score)

        # Размер области графика
        drawing = Drawing(400, 200)

        chart = HorizontalLineChart()
        chart.x = 50
        chart.y = 40
        chart.height = 120
        chart.width = 300

        chart.data = [scores]
        chart.categoryAxis.categoryNames = labels

        chart.lines[0].strokeWidth = 2

        drawing.add(chart)

        renderPDF.draw(drawing, c, 40, y - 160)
        y -= 180  # отступ после графика

        line("")  # пустая строка после графика
        
    # Данные из Excel (можно всё — будет длинно)
    if hasattr(patient, "data") and isinstance(patient.data, dict):
        line("All data from Excel:", step=18)
        for k, v in patient.data.items():
            line(f"{k}: {v}")
        line("")

        


    # Если обследований нет
    if not assessments.exists():
        line("No assessments yet.")
        c.showPage()
        c.save()
        return response

    # Сводка по обследованиям
    line("Assessments summary:", step=18)
    for a in assessments:
        dt = timezone.localtime(a.created_at).strftime("%d.%m.%Y %H:%M")
        line(f"- {dt} | {a.questionnaire.title} | score={a.total_score} | {a.conclusion}")
    line("")

    # Детально: ответы по каждому обследованию
    for a in assessments:
        line("=" * 90)
        dt = timezone.localtime(a.created_at).strftime("%d.%m.%Y %H:%M")
        line(f"Assessment #{a.id} | {dt}", step=18)
        line(f"Questionnaire: {a.questionnaire.title}")
        line(f"Total score: {a.total_score}")
        line(f"Conclusion: {a.conclusion}")
        line("")

        answers = (
            Answer.objects
            .filter(assessment=a)
            .select_related("question")
            .order_by("question__order")
        )

        line("Answers:", step=18)
        for ans in answers:
            q = ans.question
            line(f"{q.order}. {q.text}")
            line(f"   Answer: {ans.value}   Score: {ans.score}")
        line("")

    c.showPage()
    c.save()
    return response




def export_lab_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lab Template"

    minerals = LabIndicator.objects.filter(category="mineral")
    hormones = LabIndicator.objects.filter(category="hormone")
    vitamins = LabIndicator.objects.filter(category="vitamin")

    headers = ["Patient_ID", "Date"]

    # Добавляем заголовки по группам
    headers += ["--- MINERALS ---"]
    headers += [m.name for m in minerals]

    headers += ["--- HORMONES ---"]
    headers += [h.name for h in hormones]

    headers += ["--- VITAMINS ---"]
    headers += [v.name for v in vitamins]

    ws.append(headers)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="lab_template.xlsx"'

    wb.save(response)
    return response

@require_POST
@login_required
def import_lab_excel(request):
    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Файл не выбран")
        return redirect("patient_list")

    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        messages.error(request, "В файле нет данных (только заголовки или пусто)")
        return redirect("patient_list")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    # Индексы базовых колонок
    try:
        patient_col = headers.index("Patient_ID")
        date_col = headers.index("Date")
    except ValueError:
        messages.error(request, "Неверный шаблон: нет колонок Patient_ID/Date")
        return redirect("patient_list")

    # Подготовим карту: index колонки -> LabIndicator
    indicator_by_col = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        if h.startswith("---"):  # разделители групп
            continue
        if h in ("Patient_ID", "Date"):

            continue

        ind = LabIndicator.objects.filter(name=h).first()
        if ind:
            indicator_by_col[i] = ind

    imported_results = 0
    imported_values = 0

    for r in rows[1:]:
        if not r or all(v is None or str(v).strip() == "" for v in r):
         continue

        patient_code = r[patient_col]
        dt_val = r[date_col]

        if not patient_code:
            continue

        try:
            patient = Patient.objects.get(patient_code=str(patient_code).strip())
        except:
        # если пациент не найден — пропускаем строку
            continue
        # Парсим дату
        parsed_date = None
        if isinstance(dt_val, datetime):
            parsed_date = dt_val.date()
        elif isinstance(dt_val, date):
            parsed_date = dt_val
        elif isinstance(dt_val, str) and dt_val.strip():
            # ожидаем формат dd.mm.yyyy или yyyy-mm-dd
            s = dt_val.strip()
            try:
                parsed_date = datetime.strptime(s, "%d.%m.%Y").date()
            except ValueError:
                try:
                    parsed_date = datetime.strptime(s, "%Y-%m-%d").date()
                except ValueError:
                    parsed_date = None

        if not parsed_date:
    # если дата пустая/кривая — ставим текущую дату Алматы
            parsed_date = timezone.localdate()

        # Создаем "шапку" результата
        lab_result = LabResult.objects.create(patient=patient, date=parsed_date)
        imported_results += 1

        # Значения
        for col_idx, indicator in indicator_by_col.items():
            if col_idx >= len(r):
                continue
            val = r[col_idx]
            if val is None or str(val).strip() == "":
                continue
            try:
                num = float(val)
            except Exception:
                continue

            LabValue.objects.create(result=lab_result, indicator=indicator, value=num)
            imported_values += 1

    messages.success(request, f"Импортировано: результатов={imported_results}, значений={imported_values}")
    return redirect("patient_list")

@require_POST
@login_required
def import_patients_excel(request):
    file = request.FILES.get("file")
    if not file:
        messages.error(request, "Файл не выбран")
        return redirect("patient_list")

    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        messages.error(request, "В файле нет данных")
        return redirect("patient_list")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    def get_col(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    code_col = get_col("Patient_ID")
    age_col = get_col("age")
    sex_col = get_col("sex 1-men.2-women")
    height_col = get_col("h(sm)")
    weight_col = get_col("m")

    if code_col is None:
        messages.error(request, "В файле нет колонки Patient_ID")
        return redirect("patient_list")

    created = 0
    skipped = 0

    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        patient_code = row[code_col]
        if patient_code is None:
            skipped += 1
            continue

        patient_code = str(patient_code).strip()

        # если уже существует — пропускаем
        if Patient.objects.filter(patient_code=patient_code).exists():
            skipped += 1
            continue

        age = row[age_col] if age_col is not None else None
        sex = row[sex_col] if sex_col is not None else None
        height_cm = row[height_col] if height_col is not None else None
        weight_kg = row[weight_col] if weight_col is not None else None

        # сохраняем все колонки в data
        data = {}
        for i, key in enumerate(headers):
            if key:
                data[key] = row[i] if i < len(row) else None

        sex_txt = "М" if sex == 1 else ("Ж" if sex == 2 else "")
        full_name = f"Пациент {patient_code} {sex_txt} {age if age is not None else ''}".strip()

        Patient.objects.create(
            patient_code=patient_code,
            full_name=full_name,
            age=int(age) if age is not None else None,
            sex=int(sex) if sex is not None else None,
            height_cm=float(height_cm) if height_cm is not None else None,
            weight_kg=float(weight_kg) if weight_kg is not None else None,
            data=data,
        )
        created += 1

    messages.success(request, f"Импорт завершён: добавлено {created}, пропущено {skipped}")
    return redirect("patient_list")


@require_POST
@login_required
def delete_patient(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    patient.delete()
    messages.success(request, "Пациент удалён")
    return redirect("patient_list")


@require_POST
@login_required
def add_patient_note(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    text = request.POST.get("text", "").strip()

    if text:
        PatientNote.objects.create(
            patient=patient,
            doctor=request.user,
            text=text,
        )
        messages.success(request, "Заметка добавлена")
    else:
        messages.error(request, "Текст заметки пустой")

    return redirect("patient_detail", patient_id=patient.id)

@require_POST
@login_required
def delete_patient_note(request, note_id):
    note = get_object_or_404(PatientNote, id=note_id)

    # только автор может удалить
    if note.doctor != request.user:
        messages.error(request, "Нет прав для удаления")
        return redirect("patient_detail", patient_id=note.patient.id)

    patient_id = note.patient.id
    note.delete()

    messages.success(request, "Заметка удалена")
    return redirect("patient_detail", patient_id=patient_id)

@login_required
def edit_patient_note(request, note_id):
    note = get_object_or_404(PatientNote, id=note_id)

    # только автор может редактировать
    if note.doctor != request.user:
        messages.error(request, "Нет прав для редактирования")
        return redirect("patient_detail", patient_id=note.patient.id)

    if request.method == "POST":
        text = request.POST.get("text", "").strip()

        if text:
            note.text = text
            note.save()
            messages.success(request, "Заметка обновлена")
            return redirect("patient_detail", patient_id=note.patient.id)
        else:
            messages.error(request, "Текст пустой")

    return render(request, "core/edit_note.html", {"note": note})
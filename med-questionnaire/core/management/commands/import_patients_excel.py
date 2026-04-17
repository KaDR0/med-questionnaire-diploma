from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from core.models import Patient


class Command(BaseCommand):
    help = "Import patients from Excel"

    def add_arguments(self, parser):
        parser.add_argument("filepath", type=str, help="Path to xlsx file")
        parser.add_argument("--sheet", type=str, default="Лист1", help="Sheet name")
        parser.add_argument("--limit", type=int, default=0, help="How many rows to import (0=all)")

    def handle(self, *args, **options):
        filepath = options["filepath"]
        sheet = options["sheet"]
        limit = options["limit"]

        wb = load_workbook(filepath, data_only=True)
        ws = wb[sheet]

        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        header = [str(h).strip() if h is not None else "" for h in header]

        def idx(name):
            try:
                return header.index(name)
            except ValueError:
                return None

        i_age = idx("age")
        i_sex = idx("sex 1-men.2-women")
        i_h = idx("h(sm)")
        i_w = idx("m")

        created = 0
        for n, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            if not row or all(v is None for v in row):
                continue

            # лимит
            if limit and created >= limit:
                break

            age = row[i_age] if i_age is not None else None
            sex = row[i_sex] if i_sex is not None else None
            h = row[i_h] if i_h is not None else None
            w = row[i_w] if i_w is not None else None

            # сохраняем ВСЕ колонки в data
            data = {}
            for i, key in enumerate(header):
                if key:
                    data[key] = row[i]

            # красивое имя (чтобы в админке было удобно)
            sex_txt = "М" if sex == 1 else ("Ж" if sex == 2 else "")
            full_name = f"Пациент {created+1} {sex_txt} {age if age is not None else ''}".strip()

            Patient.objects.create(
                full_name=full_name,
                age=int(age) if age is not None else None,
                sex=int(sex) if sex is not None else None,
                height_cm=float(h) if h is not None else None,
                weight_kg=float(w) if w is not None else None,
                data=data,
            )
            created += 1

        self.stdout.write(self.style.SUCCESS(f"Imported patients: {created}"))